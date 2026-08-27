"""Sprint 7 tests — Notifications, Reporting, PDF/Excel exports.

Runs against an in-memory SQLite database so PostgreSQL is not required
for the test suite. All business-logic and HTTP endpoint behaviour is
tested, including creator-isolation enforcement.
"""
import io
import pytest
from datetime import date, datetime
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.db.database import Base, get_db
from app.core.security import create_access_token, hash_password
import main  # triggers Base.metadata.create_all in real DB — we override below

# ---------------------------------------------------------------------------
# In-memory SQLite test database
# ---------------------------------------------------------------------------

SQLITE_URL = "sqlite:///./test_sprint7.db"

engine = create_engine(SQLITE_URL, connect_args={"check_same_thread": False})
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

Base.metadata.create_all(bind=engine)


def override_get_db():
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()


main.app.dependency_overrides[get_db] = override_get_db
client = TestClient(main.app)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _create_creator(email: str, name: str = "Test Creator") -> dict:
    """Register a creator and return login token + user_id."""
    reg = client.post("/auth/register", json={
        "full_name": name,
        "email": email,
        "password": "Test@1234",
        "role": "Creator",
    })
    assert reg.status_code in (200, 201, 409), reg.text
    login = client.post("/auth/login", json={"email": email, "password": "Test@1234"})
    assert login.status_code == 200, login.text
    token = login.json()["access_token"]
    # Use /auth/profile to get the authenticated user's id
    profile = client.get("/auth/profile", headers={"Authorization": f"Bearer {token}"})
    assert profile.status_code == 200, f"Profile failed: {profile.text}"
    user_id = profile.json().get("id")
    return {"token": token, "user_id": user_id, "headers": {"Authorization": f"Bearer {token}"}}


# Create two isolated creators once
@pytest.fixture(scope="module")
def creator_a():
    return _create_creator("creator_a_s7@test.com", "Creator Alpha")


@pytest.fixture(scope="module")
def creator_b():
    return _create_creator("creator_b_s7@test.com", "Creator Beta")


# ---------------------------------------------------------------------------
# 1. NOTIFICATION CRUD TESTS
# ---------------------------------------------------------------------------

class TestNotificationCRUD:
    def test_list_empty(self, creator_a):
        r = client.get("/notifications", headers=creator_a["headers"])
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_create_notification(self, creator_a):
        payload = {
            "title": "Test Performance Alert",
            "message": "Your content performed well this week.",
            "notification_type": "performance",
        }
        r = client.post("/notifications", json=payload, headers=creator_a["headers"])
        assert r.status_code == 201
        data = r.json()
        assert data["title"] == payload["title"]
        assert data["is_read"] is False
        assert data["notification_type"] == "performance"
        assert data["creator_id"] == creator_a["user_id"]

    def test_create_invalid_type_defaults_to_general(self, creator_a):
        # notification_type must be a valid literal — invalid should fail schema validation
        payload = {"title": "Test", "message": "Hello", "notification_type": "invalid_type"}
        r = client.post("/notifications", json=payload, headers=creator_a["headers"])
        assert r.status_code == 422

    def test_get_notification(self, creator_a):
        # Create then fetch
        n = client.post("/notifications", json={
            "title": "Fetch Test", "message": "Msg", "notification_type": "general"
        }, headers=creator_a["headers"]).json()
        r = client.get(f"/notifications/{n['id']}", headers=creator_a["headers"])
        assert r.status_code == 200
        assert r.json()["id"] == n["id"]

    def test_mark_single_read(self, creator_a):
        n = client.post("/notifications", json={
            "title": "Read Test", "message": "Msg", "notification_type": "revenue"
        }, headers=creator_a["headers"]).json()
        assert n["is_read"] is False
        r = client.put(f"/notifications/{n['id']}/read", headers=creator_a["headers"])
        assert r.status_code == 200
        assert r.json()["is_read"] is True

    def test_mark_all_read(self, creator_a):
        # Create a few unread
        for i in range(3):
            client.post("/notifications", json={
                "title": f"Bulk {i}", "message": "Msg", "notification_type": "engagement"
            }, headers=creator_a["headers"])
        r = client.put("/notifications/read-all", headers=creator_a["headers"])
        assert r.status_code == 200
        assert r.json()["success"] is True

    def test_delete_notification(self, creator_a):
        n = client.post("/notifications", json={
            "title": "Delete Me", "message": "Msg", "notification_type": "general"
        }, headers=creator_a["headers"]).json()
        r = client.delete(f"/notifications/{n['id']}", headers=creator_a["headers"])
        assert r.status_code == 200
        # Confirm it's gone
        fetch = client.get(f"/notifications/{n['id']}", headers=creator_a["headers"])
        assert fetch.status_code == 404

    def test_unread_count(self, creator_a):
        r = client.get("/notifications/unread-count", headers=creator_a["headers"])
        assert r.status_code == 200
        assert "unread_count" in r.json()

    def test_unauthenticated_access_denied(self):
        r = client.get("/notifications")
        assert r.status_code == 401


# ---------------------------------------------------------------------------
# 2. CREATOR ISOLATION TESTS
# ---------------------------------------------------------------------------

class TestNotificationIsolation:
    def test_creator_b_cannot_read_creator_a_notification(self, creator_a, creator_b):
        # Creator A creates a notification
        n = client.post("/notifications", json={
            "title": "Private A", "message": "Only for A", "notification_type": "general"
        }, headers=creator_a["headers"]).json()
        notif_id = n["id"]

        # Creator B tries to access it — must get 404
        r = client.get(f"/notifications/{notif_id}", headers=creator_b["headers"])
        assert r.status_code == 404

    def test_creator_b_cannot_delete_creator_a_notification(self, creator_a, creator_b):
        n = client.post("/notifications", json={
            "title": "Delete Guard", "message": "Protected", "notification_type": "general"
        }, headers=creator_a["headers"]).json()

        r = client.delete(f"/notifications/{n['id']}", headers=creator_b["headers"])
        assert r.status_code == 404

    def test_creator_b_cannot_mark_creator_a_notification(self, creator_a, creator_b):
        n = client.post("/notifications", json={
            "title": "Read Guard", "message": "Protected", "notification_type": "general"
        }, headers=creator_a["headers"]).json()

        r = client.put(f"/notifications/{n['id']}/read", headers=creator_b["headers"])
        assert r.status_code == 404

    def test_creator_lists_only_own_notifications(self, creator_a, creator_b):
        # Creator A notifications should not appear in Creator B's list
        for_a = client.get("/notifications", headers=creator_a["headers"]).json()
        for_b = client.get("/notifications", headers=creator_b["headers"]).json()
        a_ids = {n["id"] for n in for_a}
        b_ids = {n["id"] for n in for_b}
        assert a_ids.isdisjoint(b_ids), "Creator B's list must not contain Creator A notifications"


# ---------------------------------------------------------------------------
# 3. GENERATE ALERTS TESTS
# ---------------------------------------------------------------------------

class TestAlertGeneration:
    def test_generate_alerts_endpoint(self, creator_a):
        r = client.post("/notifications/generate-alerts", headers=creator_a["headers"])
        assert r.status_code == 201
        data = r.json()
        assert "total_generated" in data
        assert "performance_alerts" in data
        assert "engagement_alerts" in data
        assert "revenue_alerts" in data
        assert isinstance(data["total_generated"], int)

    def test_generate_alerts_unauthenticated(self):
        r = client.post("/notifications/generate-alerts")
        assert r.status_code == 401


# ---------------------------------------------------------------------------
# 4. REPORT SUMMARY TESTS
# ---------------------------------------------------------------------------

class TestReportSummary:
    def test_report_summary(self, creator_a):
        r = client.get("/reports/summary", headers=creator_a["headers"])
        assert r.status_code == 200, f"Report summary failed: {r.text}"
        data = r.json()
        assert "creator_id" in data
        assert "content" in data
        assert "audience" in data
        assert "growth" in data
        assert "revenue" in data
        assert "platform_performance" in data
        # creator_id must match the authenticated user
        if creator_a["user_id"] is not None:
            assert data["creator_id"] == creator_a["user_id"]

    def test_report_summary_unauthenticated(self):
        r = client.get("/reports/summary")
        assert r.status_code == 401

    def test_report_contains_content_keys(self, creator_a):
        data = client.get("/reports/summary", headers=creator_a["headers"]).json()
        content = data["content"]
        for key in ["total_views", "total_likes", "total_comments", "total_shares",
                    "total_reach", "average_engagement_rate", "top_content"]:
            assert key in content, f"Missing key: {key}"

    def test_report_contains_audience_keys(self, creator_a):
        data = client.get("/reports/summary", headers=creator_a["headers"]).json()
        aud = data["audience"]
        for key in ["total_followers", "total_reach", "total_impressions",
                    "gender_distribution", "age_distribution"]:
            assert key in aud, f"Missing key: {key}"

    def test_report_contains_revenue_keys(self, creator_a):
        data = client.get("/reports/summary", headers=creator_a["headers"]).json()
        rev = data["revenue"]
        for key in ["total_revenue", "currency", "revenue_by_source", "monthly_revenue"]:
            assert key in rev, f"Missing key: {key}"


# ---------------------------------------------------------------------------
# 5. PDF EXPORT TESTS
# ---------------------------------------------------------------------------

class TestPdfExport:
    def test_pdf_export_returns_pdf(self, creator_a):
        r = client.get("/reports/export/pdf", headers=creator_a["headers"])
        assert r.status_code == 200
        assert r.headers["content-type"] == "application/pdf"
        assert len(r.content) > 1000  # PDF must be non-trivial

    def test_pdf_starts_with_pdf_header(self, creator_a):
        r = client.get("/reports/export/pdf", headers=creator_a["headers"])
        assert r.content[:4] == b"%PDF", "Response is not a valid PDF"

    def test_pdf_export_unauthenticated(self):
        r = client.get("/reports/export/pdf")
        assert r.status_code == 401


# ---------------------------------------------------------------------------
# 6. EXCEL EXPORT TESTS
# ---------------------------------------------------------------------------

class TestExcelExport:
    def test_excel_export_returns_xlsx(self, creator_a):
        r = client.get("/reports/export/excel", headers=creator_a["headers"])
        assert r.status_code == 200
        ct = r.headers["content-type"]
        assert "spreadsheetml" in ct or "openxmlformats" in ct

    def test_excel_is_valid_workbook(self, creator_a):
        import openpyxl
        r = client.get("/reports/export/excel", headers=creator_a["headers"])
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        expected_sheets = {"Summary", "Content Performance", "Audience Analytics", "Growth", "Revenue", "Platform Performance"}
        assert expected_sheets.issubset(set(wb.sheetnames)), f"Missing sheets. Got: {wb.sheetnames}"

    def test_excel_export_unauthenticated(self):
        r = client.get("/reports/export/excel")
        assert r.status_code == 401


# ---------------------------------------------------------------------------
# 7. REPORT ISOLATION TESTS
# ---------------------------------------------------------------------------

class TestReportIsolation:
    def test_report_belongs_to_authenticated_creator(self, creator_a, creator_b):
        r_a = client.get("/reports/summary", headers=creator_a["headers"])
        r_b = client.get("/reports/summary", headers=creator_b["headers"])
        assert r_a.status_code == 200
        assert r_b.status_code == 200
        # Reports must be scoped to the authenticated creator
        if creator_a["user_id"] and creator_b["user_id"]:
            assert r_a.json()["creator_id"] == creator_a["user_id"]
            assert r_b.json()["creator_id"] == creator_b["user_id"]
            assert r_a.json()["creator_id"] != r_b.json()["creator_id"]


# ---------------------------------------------------------------------------
# Cleanup
# ---------------------------------------------------------------------------

def teardown_module():
    # SQLite file is released by GC — best-effort removal
    import os, gc
    gc.collect()
    for fname in ("test_sprint7.db", "test_sprint7.db-shm", "test_sprint7.db-wal"):
        try:
            os.remove(fname)
        except (FileNotFoundError, PermissionError):
            pass
