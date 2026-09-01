"""
CreatorIQ Report Service

Builds structured reports from existing analytics, audience,
revenue, and sponsorship modules (no duplicated analytics logic).
Supports JSON payload + Excel/PDF export.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from sqlalchemy.orm import Session

from app.models.content import Content
from app.models.user import User

from app.services.analytics_service import (
    calculate_engagement_rate,
    get_kpi_summary,
    get_platform_comparison,
    get_platform_performance,
    get_top_content,
)

from app.services.audience_service import (
    get_audience_report,
    get_growth_report,
)

from app.services import revenue_service
from app.services.sponsorship_service import get_all_sponsorships


# ============================================================
# REPORT TYPE CATALOGUE
# ============================================================

REPORT_TYPES = [
    {
        "key": "full",
        "name": "Executive Comprehensive Report",
        "description": (
            "All-in-one executive analysis combining Content, "
            "Audience, Revenue, and Growth metrics."
        ),
    },
    {
        "key": "content",
        "name": "Content Performance Report",
        "description": (
            "Views, likes, shares, comments, and engagement "
            "rates by content and platform."
        ),
    },
    {
        "key": "audience",
        "name": "Audience Analytics Report",
        "description": (
            "Demographics including age groups, devices, "
            "locations, and reach trends."
        ),
    },
    {
        "key": "revenue",
        "name": "Revenue Analytics Report",
        "description": (
            "Revenue streams, monthly earnings, and "
            "sponsorship deal statuses."
        ),
    },
    {
        "key": "growth",
        "name": "Growth Trends Report",
        "description": (
            "Follower growth, engagement trajectory, "
            "and trend analysis."
        ),
    },
    {
        "key": "platform",
        "name": "Platform Comparison Report",
        "description": (
            "Cross-platform analytics comparing YouTube, "
            "Instagram, and other connected platforms."
        ),
    },
]

# Accept reference-style aliases from other templates
_TYPE_ALIASES = {
    "executive_summary": "full",
    "content_performance": "content",
    "audience_analytics": "audience",
    "revenue_analytics": "revenue",
    "growth_trends": "growth",
    "platform_comparison": "platform",
}


def get_available_report_types() -> list[dict[str, str]]:
    return list(REPORT_TYPES)


def _normalize_report_type(report_type: str | None) -> str:
    key = (report_type or "full").lower().strip()
    key = _TYPE_ALIASES.get(key, key)
    allowed = {r["key"] for r in REPORT_TYPES}
    if key not in allowed:
        raise ValueError(
            "Invalid report_type. Use: "
            + ", ".join(sorted(allowed))
        )
    return key


def _type_meta(report_type: str) -> dict[str, str]:
    return next(
        (r for r in REPORT_TYPES if r["key"] == report_type),
        REPORT_TYPES[0],
    )


def _serialize_sponsorships(sponsorships) -> list[dict]:
    rows = []
    for s in sponsorships:
        rows.append(
            {
                "id": s.id,
                "brand_name": s.brand_name,
                "campaign_name": s.campaign_name,
                "contract_value": getattr(s, "contract_value", 0.0) or 0.0,
                "status": s.status,
                "payment_status": s.payment_status,
                "start_date": s.start_date.isoformat() if s.start_date else None,
                "end_date": s.end_date.isoformat() if s.end_date else None,
            }
        )
    return rows


def _build_insights_and_recommendations(
    kpis: dict,
    revenue_by_source: list,
    sponsorships: list,
) -> tuple[list[str], list[str]]:
    insights: list[str] = []
    recommendations: list[str] = []

    if kpis.get("best_platform"):
        insights.append(
            f"Highest performing channel is {kpis['best_platform']} "
            "with strong relative engagement."
        )

    total_revenue = float(kpis.get("total_revenue") or 0)
    if total_revenue > 0:
        insights.append(
            f"Recorded cumulative earnings total ${total_revenue:,.2f}."
        )
        if revenue_by_source:
            # support list of dicts with source/amount or total
            def amount_of(row):
                if isinstance(row, dict):
                    return float(row.get("amount") or row.get("total") or 0)
                return 0.0

            top = max(revenue_by_source, key=amount_of)
            if isinstance(top, dict):
                src = top.get("source") or top.get("revenue_source") or "unknown"
                insights.append(
                    f"Primary revenue driver is '{src}' "
                    f"(${amount_of(top):,.2f})."
                )

    eng = float(kpis.get("average_engagement_rate") or 0)
    if eng >= 5.0:
        recommendations.append(
            "Audience engagement is strong (≥5%). "
            "Leverage current formats for brand deals."
        )
    else:
        recommendations.append(
            "To lift engagement toward 5%+, optimize posting "
            "schedules and strengthen calls-to-action."
        )

    if sponsorships:
        pending = [
            s
            for s in sponsorships
            if str(getattr(s, "payment_status", "") or "").lower()
            in {"pending", "unpaid", "partial"}
        ]
        if pending:
            recommendations.append(
                f"Follow up on {len(pending)} sponsorship payment(s) "
                "still pending or partial."
            )

    if not insights:
        insights.append(
            "Continue syncing platform data so KPIs stay current."
        )

    return insights, recommendations


# ============================================================
# BUILD REPORT (used by routers/reports.py)
# ============================================================

def build_creator_report(
    db: Session,
    creator_id: int | None = None,
    report_type: str = "full",
) -> dict[str, Any]:
    """
    Gather real data from existing services into one structured report.

    creator_id:
        Scope to one creator. None = all creators (admin-style).
    report_type:
        full | content | audience | revenue | growth | platform
        (also accepts executive_summary, content_performance, ...)
    """
    report_type = _normalize_report_type(report_type)
    meta = _type_meta(report_type)
    generated_at = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")

    # ---- creator profile ----
    creator_name = "All creators"
    creator_email = None
    if creator_id is not None:
        user = db.query(User).filter(User.id == creator_id).first()
        creator_name = (
            (user.full_name if user and getattr(user, "full_name", None) else None)
            or (user.email if user else None)
            or f"Creator {creator_id}"
        )
        creator_email = user.email if user else None

    # ---- content ----
    content_query = db.query(Content)
    if creator_id is not None:
        content_query = content_query.filter(Content.creator_id == creator_id)
    contents = content_query.all()

    content_performance: list[dict] = []
    c_views = c_likes = c_comments = c_shares = c_saves = c_reach = 0
    rates: list[float] = []

    for content in contents:
        total_engagement, engagement_rate = calculate_engagement_rate(content)
        c_views += content.views or 0
        c_likes += content.likes or 0
        c_comments += content.comments or 0
        c_shares += content.shares or 0
        c_saves += content.saves or 0
        c_reach += content.reach or 0
        rates.append(float(engagement_rate or 0))

        content_performance.append(
            {
                "id": content.id,
                "creator_id": content.creator_id,
                "title": content.content_title,
                "platform": content.platform,
                "views": content.views or 0,
                "likes": content.likes or 0,
                "comments": content.comments or 0,
                "shares": content.shares or 0,
                "saves": content.saves or 0,
                "reach": content.reach or 0,
                "watch_time": content.watch_time or 0,
                "total_engagement": total_engagement,
                "engagement_rate": engagement_rate,
                "published_date": (
                    content.published_date.isoformat()
                    if content.published_date
                    else None
                ),
            }
        )

    c_avg_eng = round(sum(rates) / len(rates), 2) if rates else 0.0

    # ---- existing service modules ----
    analytics_summary = get_kpi_summary(db) or {}
    top_content = get_top_content(db) or []
    platform_perf = get_platform_performance(db) or []
    platform_comparison = get_platform_comparison(db) or {}

    audience_report = get_audience_report(db, creator_id) or {}
    growth_report = get_growth_report(db, creator_id) or {}

    revenue_summary = revenue_service.get_revenue_summary(db, creator_id) or {}
    revenue_by_source = revenue_service.get_revenue_by_source(db, creator_id) or []
    monthly_revenue = revenue_service.get_monthly_revenue(db, creator_id) or []
    try:
        revenue_trend = revenue_service.get_revenue_trend(db, creator_id) or []
    except Exception:
        revenue_trend = []

    sponsorships = []
    if creator_id is not None:
        try:
            sponsorships = get_all_sponsorships(db, creator_id) or []
        except Exception:
            sponsorships = []

    active_sponsorships = len(
        [
            s
            for s in sponsorships
            if str(getattr(s, "status", "") or "").lower()
            in {"active", "in progress", "in_progress"}
        ]
    )

    kpis = {
        "total_views": c_views or analytics_summary.get("total_views", 0),
        "total_likes": c_likes or analytics_summary.get("total_likes", 0),
        "total_comments": c_comments or analytics_summary.get("total_comments", 0),
        "total_shares": c_shares or analytics_summary.get("total_shares", 0),
        "total_saves": c_saves,
        "total_reach": c_reach or analytics_summary.get("total_reach", 0),
        "average_engagement_rate": (
            c_avg_eng
            if c_avg_eng > 0
            else analytics_summary.get("average_engagement_rate", 0.0)
        ),
        "total_followers": (
            audience_report.get("total_followers")
            or analytics_summary.get("total_followers", 0)
        ),
        "total_revenue": revenue_summary.get("total_revenue", 0.0),
        "active_sponsorships": active_sponsorships,
        "best_platform": analytics_summary.get("best_platform"),
        "total_content_items": len(contents),
    }

    insights, recommendations = _build_insights_and_recommendations(
        kpis, revenue_by_source, sponsorships
    )

    # ---- assemble sections by type ----
    report: dict[str, Any] = {
        "title": f"CreatorIQ {meta['name']}",
        "report_type": report_type,
        "report_type_name": meta["name"],
        "generated_at": generated_at,
        "scope": "creator" if creator_id is not None else "all",
        "creator_id": creator_id,
        "creator": {
            "id": creator_id,
            "name": creator_name,
            "email": creator_email,
        },
        # Compatible with older export code + frontend
        "summary": kpis,
        "kpis": kpis,
        "insights": insights,
        "recommendations": recommendations,
        "tables": {},
    }

    if report_type in {"full", "content"}:
        report["content_performance"] = content_performance
        report["tables"]["content_performance"] = content_performance
        report["tables"]["top_content"] = top_content

    if report_type in {"full", "audience"}:
        report["audience"] = audience_report
        report["tables"]["audience_demographics"] = audience_report

    if report_type in {"full", "growth"}:
        report["growth"] = growth_report
        report["tables"]["growth"] = growth_report

    if report_type in {"full", "revenue"}:
        report["revenue"] = {
            "summary": revenue_summary,
            "by_source": revenue_by_source,
            "monthly": monthly_revenue,
            "trend": revenue_trend,
        }
        report["tables"]["revenue_by_source"] = revenue_by_source
        report["tables"]["monthly_revenue"] = monthly_revenue
        report["tables"]["sponsorships"] = _serialize_sponsorships(sponsorships)

    if report_type in {"full", "platform"}:
        report["platform_comparison"] = platform_comparison
        report["platform_performance"] = platform_perf
        report["tables"]["platform_performance"] = platform_perf
        report["tables"]["platform_comparison"] = platform_comparison

    return report


# Alias matching the reference naming
generate_report_data = build_creator_report


# ============================================================
# EXCEL / PDF EXPORT (unchanged behaviour for routers)
# ============================================================

def _append_dict_section(worksheet, title, data):
    worksheet.append([title])
    if not isinstance(data, dict):
        worksheet.append(["Value", data])
        worksheet.append([])
        return
    worksheet.append(["Metric", "Value"])
    for key, value in data.items():
        if isinstance(value, (dict, list)):
            continue
        worksheet.append([key, value])
    worksheet.append([])


def _append_list_section(worksheet, title, rows):
    worksheet.append([title])
    if not rows:
        worksheet.append(["No data"])
        worksheet.append([])
        return
    if isinstance(rows, dict):
        _append_dict_section(worksheet, title, rows)
        return
    if not isinstance(rows, list):
        worksheet.append([str(rows)])
        worksheet.append([])
        return
    first = rows[0]
    if isinstance(first, dict):
        headers = list(first.keys())
        worksheet.append(headers)
        for row in rows:
            worksheet.append([row.get(h) for h in headers])
    else:
        for row in rows:
            worksheet.append([row])
    worksheet.append([])



def export_report_excel(report: dict) -> bytes:
    """
    Colored multi-sheet Excel export with creator details,
    KPIs, insights, content, audience, revenue, platforms.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for Excel export. Run: pip install openpyxl"
        ) from exc

    wb = Workbook()

    # ---- styles ----
    header_fill = PatternFill("solid", fgColor="0EA5E9")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=16, color="0F172A")
    section_font = Font(bold=True, size=12, color="0369A1")
    label_fill = PatternFill("solid", fgColor="E0F2FE")
    label_font = Font(bold=True, color="0F172A", size=10)
    zebra = PatternFill("solid", fgColor="F8FAFC")
    insight_fill = PatternFill("solid", fgColor="ECFDF5")
    rec_fill = PatternFill("solid", fgColor="FEF3C7")
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    center = Alignment(vertical="center", wrap_text=True)

    def style_header_row(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin

    def autosize(ws):
        for column_cells in ws.columns:
            max_len = 0
            col = column_cells[0].column
            for cell in column_cells:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 42)

    def write_kv_table(ws, start_row, title, data: dict):
        r = start_row
        ws.cell(row=r, column=1, value=title).font = section_font
        r += 1
        ws.cell(row=r, column=1, value="Metric").font = header_font
        ws.cell(row=r, column=1).fill = header_fill
        ws.cell(row=r, column=2, value="Value").font = header_font
        ws.cell(row=r, column=2).fill = header_fill
        for c in (1, 2):
            ws.cell(row=r, column=c).border = thin
        r += 1
        if not isinstance(data, dict):
            ws.cell(row=r, column=1, value="Value")
            ws.cell(row=r, column=2, value=str(data))
            return r + 2
        i = 0
        for k, v in data.items():
            if isinstance(v, (dict, list)):
                continue
            ws.cell(row=r, column=1, value=str(k).replace("_", " ").title()).font = label_font
            ws.cell(row=r, column=1).fill = label_fill
            ws.cell(row=r, column=2, value=v if v is not None else "—")
            for c in (1, 2):
                ws.cell(row=r, column=c).border = thin
                if i % 2 == 1:
                    if c == 2:
                        ws.cell(row=r, column=c).fill = zebra
            r += 1
            i += 1
        return r + 1

    def write_list_table(ws, start_row, title, rows):
        r = start_row
        ws.cell(row=r, column=1, value=title).font = section_font
        r += 1
        if not rows:
            ws.cell(row=r, column=1, value="No data")
            return r + 2
        if isinstance(rows, dict):
            return write_kv_table(ws, start_row, title, rows)
        if not isinstance(rows, list):
            ws.cell(row=r, column=1, value=str(rows))
            return r + 2
        first = rows[0]
        if not isinstance(first, dict):
            for item in rows:
                ws.cell(row=r, column=1, value=str(item))
                r += 1
            return r + 1
        headers = list(first.keys())
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=ci, value=str(h).replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin
            cell.alignment = center
        r += 1
        for i, row in enumerate(rows):
            for ci, h in enumerate(headers, 1):
                val = row.get(h)
                if isinstance(val, (dict, list)):
                    val = str(val)
                cell = ws.cell(row=r, column=ci, value=val if val is not None else "—")
                cell.border = thin
                if i % 2 == 1:
                    cell.fill = zebra
            r += 1
        return r + 1

    # ========== Sheet: Summary ==========
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value=report.get("title") or "CreatorIQ Report").font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    creator = report.get("creator") or {}
    if not isinstance(creator, dict):
        creator = {}

    # Creator block
    ws.cell(row=3, column=1, value="Creator details").font = section_font
    creator_rows = [
        ("Creator ID", report.get("creator_id") or creator.get("id") or "—"),
        ("Creator name", creator.get("name") or "—"),
        ("Creator email", creator.get("email") or "—"),
        ("Report type", report.get("report_type_name") or report.get("report_type")),
        ("Generated at", report.get("generated_at")),
        ("Scope", report.get("scope")),
    ]
    r = 4
    for label, value in creator_rows:
        ws.cell(row=r, column=1, value=label).font = label_font
        ws.cell(row=r, column=1).fill = label_fill
        ws.cell(row=r, column=1).border = thin
        ws.cell(row=r, column=2, value=value).border = thin
        r += 1

    r += 1
    kpis = report.get("kpis") or report.get("summary") or {}
    r = write_kv_table(ws, r, "Key performance indicators", kpis)

    # Insights
    ws.cell(row=r, column=1, value="Insights").font = section_font
    r += 1
    for item in report.get("insights") or ["—"]:
        ws.cell(row=r, column=1, value=str(item)).fill = insight_fill
        ws.cell(row=r, column=1).border = thin
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Recommendations").font = section_font
    r += 1
    for item in report.get("recommendations") or ["—"]:
        ws.cell(row=r, column=1, value=str(item)).fill = rec_fill
        ws.cell(row=r, column=1).border = thin
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1

    autosize(ws)

    # ========== Content ==========
    if report.get("content_performance"):
        ws_c = wb.create_sheet("Content")
        write_list_table(ws_c, 1, "Content performance", report["content_performance"])
        autosize(ws_c)

    # ========== Audience ==========
    if report.get("audience"):
        ws_a = wb.create_sheet("Audience")
        aud = report["audience"]
        if isinstance(aud, dict):
            write_kv_table(ws_a, 1, "Audience overview", {
                k: v for k, v in aud.items() if not isinstance(v, (dict, list))
            })
            row = 14
            for key in ("gender_distribution", "age_distribution", "device_usage",
                        "top_countries", "top_cities"):
                if key in aud and aud[key]:
                    data = aud[key]
                    if isinstance(data, dict):
                        data = [{"name": k, "value": v} for k, v in data.items()]
                    row = write_list_table(ws_a, row, key.replace("_", " ").title(), data)
        autosize(ws_a)

    # ========== Growth ==========
    if report.get("growth"):
        ws_g = wb.create_sheet("Growth")
        g = report["growth"]
        if isinstance(g, dict):
            write_kv_table(ws_g, 1, "Growth overview", {
                k: v for k, v in g.items() if not isinstance(v, (dict, list))
            })
            # optional list fields
            for key, val in g.items():
                if isinstance(val, list) and val:
                    write_list_table(ws_g, 20, key.replace("_", " ").title(), val)
                    break
        elif isinstance(g, list):
            write_list_table(ws_g, 1, "Growth", g)
        autosize(ws_g)

    # ========== Revenue ==========
    if report.get("revenue"):
        ws_r = wb.create_sheet("Revenue")
        rev = report["revenue"] or {}
        row = 1
        if isinstance(rev, dict):
            if rev.get("summary"):
                row = write_kv_table(ws_r, row, "Revenue summary", rev["summary"])
            if rev.get("by_source"):
                row = write_list_table(ws_r, row, "Revenue by source", rev["by_source"])
            if rev.get("monthly"):
                row = write_list_table(ws_r, row, "Monthly revenue", rev["monthly"])
        tables = report.get("tables") or {}
        if tables.get("sponsorships"):
            write_list_table(ws_r, row + 1, "Sponsorships", tables["sponsorships"])
        autosize(ws_r)

    # ========== Platforms ==========
    if report.get("platform_performance") or report.get("platform_comparison"):
        ws_p = wb.create_sheet("Platforms")
        row = 1
        if report.get("platform_performance"):
            row = write_list_table(
                ws_p, row, "Platform performance", report["platform_performance"]
            )
        pc = report.get("platform_comparison")
        if isinstance(pc, dict):
            write_kv_table(ws_p, row, "Platform comparison", {
                k: v for k, v in pc.items() if not isinstance(v, (dict, list))
            })
        elif isinstance(pc, list):
            write_list_table(ws_p, row, "Platform comparison", pc)
        autosize(ws_p)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_report_pdf(report: dict) -> bytes:
    """
    Colored PDF with creator header, KPI table, insights,
    recommendations, and content sample table.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable,
        )
    except ImportError as exc:
        raise RuntimeError(
            "reportlab is required for PDF export. Run: pip install reportlab"
        ) from exc

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CIQTitle",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#0F172A"),
        spaceAfter=6,
    )
    h2 = ParagraphStyle(
        "CIQH2",
        parent=styles["Heading2"],
        fontSize=12,
        textColor=colors.HexColor("#0369A1"),
        spaceBefore=10,
        spaceAfter=6,
    )
    body = ParagraphStyle(
        "CIQBody",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#334155"),
        leading=12,
    )

    story = []
    story.append(Paragraph(report.get("title") or "CreatorIQ Report", title_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#0EA5E9")))
    story.append(Spacer(1, 6))

    creator = report.get("creator") or {}
    if not isinstance(creator, dict):
        creator = {}

    creator_data = [
        ["Creator ID", str(report.get("creator_id") or creator.get("id") or "—")],
        ["Creator name", str(creator.get("name") or "—")],
        ["Creator email", str(creator.get("email") or "—")],
        ["Report type", str(report.get("report_type_name") or report.get("report_type") or "—")],
        ["Generated at", str(report.get("generated_at") or "—")],
        ["Scope", str(report.get("scope") or "—")],
    ]
    t = Table(creator_data, colWidths=[120, 350])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E0F2FE")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#0F172A")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    story.append(Spacer(1, 10))

    # KPIs
    story.append(Paragraph("Key performance indicators", h2))
    kpis = report.get("kpis") or report.get("summary") or {}
    kpi_rows = [["Metric", "Value"]]
    for k, v in kpis.items():
        if isinstance(v, (dict, list)):
            continue
        kpi_rows.append([str(k).replace("_", " ").title(), str(v if v is not None else "—")])
    kt = Table(kpi_rows, colWidths=[220, 250])
    kt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0EA5E9")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#E0F2FE")),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(kt)

    # Insights
    story.append(Paragraph("Insights", h2))
    for item in report.get("insights") or ["No insights available."]:
        story.append(Paragraph(f"• {item}", body))

    story.append(Paragraph("Recommendations", h2))
    for item in report.get("recommendations") or ["No recommendations available."]:
        story.append(Paragraph(f"• {item}", body))

    # Content table sample
    content_rows = report.get("content_performance") or []
    if content_rows:
        story.append(Paragraph("Content performance", h2))
        table_data = [["Title", "Platform", "Views", "Likes", "Eng %"]]
        for row in content_rows[:20]:
            title = str(row.get("title") or "Untitled")[:40]
            table_data.append([
                title,
                str(row.get("platform") or "—"),
                str(row.get("views") or 0),
                str(row.get("likes") or 0),
                str(row.get("engagement_rate") or 0),
            ])
        ct = Table(table_data, colWidths=[180, 80, 70, 70, 60])
        ct.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8B5CF6")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F3FF")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(ct)

    # Revenue by source
    rev = report.get("revenue") or {}
    by_source = rev.get("by_source") if isinstance(rev, dict) else None
    if by_source:
        story.append(Paragraph("Revenue by source", h2))
        if isinstance(by_source, list) and by_source and isinstance(by_source[0], dict):
            headers = list(by_source[0].keys())
            rd = [headers]
            for row in by_source:
                rd.append([str(row.get(h, "—")) for h in headers])
            rt = Table(rd, colWidths=[470 / max(len(headers), 1)] * len(headers))
            rt.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10B981")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#ECFDF5")]),
            ]))
            story.append(rt)

    # Platform performance
    plat = report.get("platform_performance") or []
    if isinstance(plat, list) and plat:
        story.append(Paragraph("Platform performance", h2))
        if isinstance(plat[0], dict):
            # pick useful columns
            keys = []
            for k in ("platform", "total_views", "total_likes", "total_reach", "average_engagement_rate"):
                if k in plat[0]:
                    keys.append(k)
            if not keys:
                keys = list(plat[0].keys())[:5]
            pd = [[k.replace("_", " ").title() for k in keys]]
            for row in plat:
                pd.append([str(row.get(k, "—")) for k in keys])
            pt = Table(pd, colWidths=[470 / len(keys)] * len(keys))
            pt.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F59E0B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFFBEB")]),
            ]))
            story.append(pt)

    doc.build(story)
    return buffer.getvalue()