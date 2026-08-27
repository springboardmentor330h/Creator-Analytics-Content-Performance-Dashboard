"""Reporting service — aggregates existing analytics data into structured reports.

Does NOT duplicate calculations from analytics_service, audience_service,
growth_service, or revenue_service. Only calls existing service functions and
combines the results.
"""
import io
from datetime import datetime
from typing import Any, Dict

from sqlalchemy.orm import Session

from app.models.user import User
from app.services.analytics_service import (
    get_dashboard_summary,
    get_platform_performance,
    get_top_content,
)
from app.services.audience_service import get_audience_analytics, get_growth_analytics
from app.services.revenue_service import (
    get_monthly_revenue,
    get_revenue_by_source,
    get_revenue_trend,
    get_total_revenue,
)
from app.services.sponsorship_service import get_sponsorships_summary


def get_structured_report(db: Session, user: User) -> Dict[str, Any]:
    """Build a comprehensive creator report by combining existing service calls."""
    # --- Content Performance (from analytics_service) ---
    dashboard = get_dashboard_summary(db, user)
    top_content = get_top_content(db, user)

    content_data = {
        "total_views": dashboard.get("total_views", 0),
        "total_likes": dashboard.get("total_likes", 0),
        "total_comments": dashboard.get("total_comments", 0),
        "total_shares": dashboard.get("total_shares", 0),
        "total_reach": dashboard.get("total_reach", 0),
        "average_engagement_rate": dashboard.get("average_engagement_rate", 0.0),
        "top_content": top_content,
    }

    # --- Audience Analytics (from audience_service) ---
    audience_raw = get_audience_analytics(db, user)
    audience_data = {
        "total_followers": audience_raw.get("total_followers", 0),
        "total_reach": audience_raw.get("total_reach", 0),
        "total_impressions": audience_raw.get("total_impressions", 0),
        "gender_distribution": audience_raw.get("gender_distribution", {}),
        "age_distribution": audience_raw.get("age_distribution", {}),
        "top_countries": audience_raw.get("top_countries", [])[:5],
        "top_cities": audience_raw.get("top_cities", [])[:5],
        "device_distribution": audience_raw.get("device_distribution", {}),
    }

    # --- Growth Analytics (from audience_service.get_growth_analytics) ---
    growth_points = get_growth_analytics(db, user)
    if growth_points:
        latest = growth_points[-1]
        first = growth_points[0]
        follower_growth = latest["followers"] - first["followers"]
        growth_pct = latest["growth_percentage"]
        growth_trend = "up" if follower_growth > 0 else "down" if follower_growth < 0 else "stable"
    else:
        follower_growth = 0
        growth_pct = 0.0
        growth_trend = "stable"

    growth_data = {
        "follower_growth": follower_growth,
        "growth_percentage": growth_pct,
        "growth_trend": growth_trend,
        "data_points": growth_points,
    }

    # --- Revenue Analytics (from revenue_service) ---
    rev_total = get_total_revenue(db, user)
    rev_by_source = get_revenue_by_source(db, user)
    rev_monthly = get_monthly_revenue(db, user)
    rev_trend = get_revenue_trend(db, user)
    revenue_data = {
        "total_revenue": rev_total.get("total_revenue", 0.0),
        "currency": rev_total.get("currency", "INR"),
        "revenue_by_source": rev_by_source,
        "monthly_revenue": rev_monthly,
        "revenue_trend": rev_trend,
    }

    # --- Sponsorship Summary ---
    sponsorship_data = get_sponsorships_summary(db, user)

    # --- Platform Performance ---
    platform_perf = get_platform_performance(db, user)

    return {
        "creator_id": user.id,
        "creator_name": user.full_name,
        "generated_at": datetime.utcnow().isoformat(),
        "content": content_data,
        "audience": audience_data,
        "growth": growth_data,
        "revenue": revenue_data,
        "sponsorships": sponsorship_data,
        "platform_performance": platform_perf,
    }


# ---------------------------------------------------------------------------
# PDF Export
# ---------------------------------------------------------------------------

def generate_pdf_report(report_data: Dict[str, Any]) -> bytes:
    """Generate a professional PDF report using ReportLab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    BRAND = colors.HexColor("#1e293b")
    ACCENT = colors.HexColor("#6366f1")
    LIGHT = colors.HexColor("#f8fafc")
    MID = colors.HexColor("#e2e8f0")
    TEXT = colors.HexColor("#334155")

    h1_style = ParagraphStyle("H1", parent=styles["Heading1"], textColor=BRAND, fontSize=22, spaceAfter=4, fontName="Helvetica-Bold")
    h2_style = ParagraphStyle("H2", parent=styles["Heading2"], textColor=ACCENT, fontSize=13, spaceAfter=4, spaceBefore=12, fontName="Helvetica-Bold")
    body_style = ParagraphStyle("Body", parent=styles["Normal"], textColor=TEXT, fontSize=9, spaceAfter=2, fontName="Helvetica")
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"], textColor=colors.HexColor("#94a3b8"), fontSize=8, fontName="Helvetica")
    center_style = ParagraphStyle("Center", parent=styles["Normal"], alignment=TA_CENTER, textColor=TEXT, fontSize=9, fontName="Helvetica")

    def section_table_style(header_bg=BRAND):
        return TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
            ("GRID", (0, 0), (-1, -1), 0.3, MID),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])

    story = []

    # ----- HEADER -----
    story.append(Paragraph("CreatorIQ Analytics Report", h1_style))
    story.append(Paragraph(
        f"Creator: <b>{report_data.get('creator_name', '')}</b> &nbsp;|&nbsp; "
        f"Generated: {datetime.utcnow().strftime('%B %d, %Y %H:%M UTC')}",
        sub_style
    ))
    story.append(HRFlowable(width="100%", thickness=1.5, color=ACCENT, spaceAfter=12))

    # ----- CONTENT PERFORMANCE -----
    content = report_data.get("content", {})
    story.append(Paragraph("Content Performance", h2_style))
    kpi_data = [
        ["Metric", "Value"],
        ["Total Views", f"{content.get('total_views', 0):,}"],
        ["Total Likes", f"{content.get('total_likes', 0):,}"],
        ["Total Comments", f"{content.get('total_comments', 0):,}"],
        ["Total Shares", f"{content.get('total_shares', 0):,}"],
        ["Total Reach", f"{content.get('total_reach', 0):,}"],
        ["Avg. Engagement Rate", f"{content.get('average_engagement_rate', 0):.2f}%"],
    ]
    kpi_table = Table(kpi_data, colWidths=[200, 200])
    kpi_table.setStyle(section_table_style())
    story.append(kpi_table)

    # Top Content
    top = content.get("top_content", [])
    if top:
        story.append(Spacer(1, 8))
        story.append(Paragraph("Top Performing Content", h2_style))
        tc_data = [["Title", "Platform", "Views", "Engagement Rate"]]
        for row in top[:10]:
            tc_data.append([
                str(row.get("content_title", ""))[:40],
                str(row.get("platform", "")),
                f"{row.get('views', 0):,}",
                f"{row.get('engagement_rate', 0):.2f}%",
            ])
        tc_table = Table(tc_data, colWidths=[210, 80, 80, 100])
        tc_table.setStyle(section_table_style(ACCENT))
        story.append(tc_table)

    # ----- AUDIENCE ANALYTICS -----
    audience = report_data.get("audience", {})
    story.append(Spacer(1, 10))
    story.append(Paragraph("Audience Analytics", h2_style))
    aud_data = [
        ["Metric", "Value"],
        ["Total Followers", f"{audience.get('total_followers', 0):,}"],
        ["Total Reach", f"{audience.get('total_reach', 0):,}"],
        ["Total Impressions", f"{audience.get('total_impressions', 0):,}"],
        ["Top Country", ", ".join(audience.get("top_countries", ["—"])[:3])],
        ["Top City", ", ".join(audience.get("top_cities", ["—"])[:3])],
    ]
    aud_table = Table(aud_data, colWidths=[200, 250])
    aud_table.setStyle(section_table_style())
    story.append(aud_table)

    # Gender distribution
    gd = audience.get("gender_distribution", {})
    if gd:
        story.append(Spacer(1, 6))
        story.append(Paragraph("Gender Distribution", h2_style))
        gd_data = [["Gender", "Percentage"]] + [[k, f"{v:.1f}%"] for k, v in gd.items()]
        gd_table = Table(gd_data, colWidths=[200, 200])
        gd_table.setStyle(section_table_style(ACCENT))
        story.append(gd_table)

    # ----- GROWTH ANALYTICS -----
    growth = report_data.get("growth", {})
    story.append(Spacer(1, 10))
    story.append(Paragraph("Growth Analytics", h2_style))
    gr_data = [
        ["Metric", "Value"],
        ["Follower Growth (Period)", f"+{growth.get('follower_growth', 0):,}"],
        ["Growth Percentage", f"{growth.get('growth_percentage', 0):.2f}%"],
        ["Growth Trend", growth.get("growth_trend", "stable").capitalize()],
    ]
    gr_table = Table(gr_data, colWidths=[200, 200])
    gr_table.setStyle(section_table_style())
    story.append(gr_table)

    # ----- REVENUE ANALYTICS -----
    revenue = report_data.get("revenue", {})
    currency = revenue.get("currency", "INR")
    story.append(Spacer(1, 10))
    story.append(Paragraph("Revenue Analytics", h2_style))
    rev_data = [
        ["Metric", "Value"],
        ["Total Revenue", f"{currency} {revenue.get('total_revenue', 0):,.2f}"],
    ]
    rev_table = Table(rev_data, colWidths=[200, 250])
    rev_table.setStyle(section_table_style())
    story.append(rev_table)

    # Revenue by source
    rev_source = revenue.get("revenue_by_source", {})
    if rev_source:
        story.append(Spacer(1, 6))
        story.append(Paragraph("Revenue by Source", h2_style))
        rs_data = [["Source", f"Amount ({currency})"]] + [
            [src, f"{amt:,.2f}"] for src, amt in rev_source.items()
        ]
        rs_table = Table(rs_data, colWidths=[200, 250])
        rs_table.setStyle(section_table_style(ACCENT))
        story.append(rs_table)

    # Monthly revenue
    monthly = revenue.get("monthly_revenue", [])
    if monthly:
        story.append(Spacer(1, 6))
        story.append(Paragraph("Monthly Revenue", h2_style))
        mr_data = [["Month", f"Revenue ({currency})"]] + [
            [m["month"], f"{m['revenue']:,.2f}"] for m in monthly
        ]
        mr_table = Table(mr_data, colWidths=[200, 250])
        mr_table.setStyle(section_table_style())
        story.append(mr_table)

    # ----- PLATFORM PERFORMANCE -----
    platforms = report_data.get("platform_performance", [])
    if platforms:
        story.append(Spacer(1, 10))
        story.append(Paragraph("Platform Performance", h2_style))
        pp_data = [["Platform", "Views", "Reach", "Likes", "Comments", "Engagement %"]]
        for p in platforms:
            pp_data.append([
                p.get("platform", ""),
                f"{p.get('total_views', 0):,}",
                f"{p.get('total_reach', 0):,}",
                f"{p.get('total_likes', 0):,}",
                f"{p.get('total_comments', 0):,}",
                f"{p.get('average_engagement_rate', 0):.2f}%",
            ])
        pp_table = Table(pp_data, colWidths=[80, 70, 70, 70, 70, 80])
        pp_table.setStyle(section_table_style(BRAND))
        story.append(pp_table)

    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=0.5, color=MID))
    story.append(Paragraph("Generated by CreatorIQ Analytics Platform", sub_style))

    doc.build(story)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

def generate_excel_report(report_data: Dict[str, Any]) -> bytes:
    """Generate a multi-sheet Excel workbook using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    BRAND_FILL = PatternFill("solid", fgColor="1e293b")
    ACCENT_FILL = PatternFill("solid", fgColor="6366f1")
    ALT_FILL = PatternFill("solid", fgColor="f1f5f9")
    WHITE_FILL = PatternFill("solid", fgColor="ffffff")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    BODY_FONT = Font(name="Calibri", size=9)
    TITLE_FONT = Font(name="Calibri", bold=True, size=13, color="1e293b")

    thin = Side(style="thin", color="e2e8f0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    def write_header(ws, row_data, fill=BRAND_FILL):
        ws.append(row_data)
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    def write_row(ws, row_data, alt=False):
        ws.append(row_data)
        fill = ALT_FILL if alt else WHITE_FILL
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.font = BODY_FONT
            cell.border = border

    def add_title(ws, title: str):
        ws.append([title])
        cell = ws[ws.max_row][0]
        cell.font = TITLE_FONT
        ws.append(["Generated:", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
                   "Creator:", report_data.get("creator_name", "")])
        ws.append([])

    # ---- Sheet 1: Summary ----
    ws1 = wb.active
    ws1.title = "Summary"
    add_title(ws1, "CreatorIQ — Analytics Report Summary")
    content = report_data.get("content", {})
    audience = report_data.get("audience", {})
    growth = report_data.get("growth", {})
    revenue = report_data.get("revenue", {})

    write_header(ws1, ["Category", "Metric", "Value"])
    summary_rows = [
        ("Content", "Total Views", content.get("total_views", 0)),
        ("Content", "Total Likes", content.get("total_likes", 0)),
        ("Content", "Total Comments", content.get("total_comments", 0)),
        ("Content", "Total Shares", content.get("total_shares", 0)),
        ("Content", "Total Reach", content.get("total_reach", 0)),
        ("Content", "Avg. Engagement Rate", f"{content.get('average_engagement_rate', 0):.2f}%"),
        ("Audience", "Total Followers", audience.get("total_followers", 0)),
        ("Audience", "Total Reach", audience.get("total_reach", 0)),
        ("Audience", "Total Impressions", audience.get("total_impressions", 0)),
        ("Growth", "Follower Growth", growth.get("follower_growth", 0)),
        ("Growth", "Growth %", f"{growth.get('growth_percentage', 0):.2f}%"),
        ("Growth", "Trend", growth.get("growth_trend", "stable")),
        ("Revenue", "Total Revenue", f"{revenue.get('currency', 'INR')} {revenue.get('total_revenue', 0):,.2f}"),
    ]
    for i, row in enumerate(summary_rows):
        write_row(ws1, list(row), alt=(i % 2 == 1))
    auto_width(ws1)

    # ---- Sheet 2: Content Performance ----
    ws2 = wb.create_sheet("Content Performance")
    add_title(ws2, "Content Performance")
    write_header(ws2, ["Title", "Platform", "Views", "Engagement Rate (%)"])
    for i, item in enumerate(content.get("top_content", [])):
        write_row(ws2, [
            item.get("content_title", ""),
            item.get("platform", ""),
            item.get("views", 0),
            round(item.get("engagement_rate", 0), 2),
        ], alt=(i % 2 == 1))
    auto_width(ws2)

    # ---- Sheet 3: Audience Analytics ----
    ws3 = wb.create_sheet("Audience Analytics")
    add_title(ws3, "Audience Analytics")
    write_header(ws3, ["Metric", "Value"])
    aud_rows = [
        ("Total Followers", audience.get("total_followers", 0)),
        ("Total Reach", audience.get("total_reach", 0)),
        ("Total Impressions", audience.get("total_impressions", 0)),
        ("Top Countries", ", ".join(audience.get("top_countries", [])[:5])),
        ("Top Cities", ", ".join(audience.get("top_cities", [])[:5])),
    ]
    for i, row in enumerate(aud_rows):
        write_row(ws3, list(row), alt=(i % 2 == 1))
    ws3.append([])
    write_header(ws3, ["Gender", "Percentage (%)"], fill=ACCENT_FILL)
    for i, (g, pct) in enumerate(audience.get("gender_distribution", {}).items()):
        write_row(ws3, [g, round(pct, 2)], alt=(i % 2 == 1))
    ws3.append([])
    write_header(ws3, ["Age Group", "Percentage (%)"], fill=ACCENT_FILL)
    for i, (age, pct) in enumerate(audience.get("age_distribution", {}).items()):
        write_row(ws3, [age, round(pct, 2)], alt=(i % 2 == 1))
    auto_width(ws3)

    # ---- Sheet 4: Growth ----
    ws4 = wb.create_sheet("Growth")
    add_title(ws4, "Growth Analytics")
    write_header(ws4, ["Date", "Followers", "Daily Growth", "Growth %"])
    for i, pt in enumerate(growth.get("data_points", [])):
        write_row(ws4, [
            pt.get("date", ""),
            pt.get("followers", 0),
            pt.get("daily_growth", 0),
            round(pt.get("growth_percentage", 0), 2),
        ], alt=(i % 2 == 1))
    auto_width(ws4)

    # ---- Sheet 5: Revenue ----
    ws5 = wb.create_sheet("Revenue")
    add_title(ws5, "Revenue Analytics")
    currency = revenue.get("currency", "INR")
    write_header(ws5, [f"Month", f"Revenue ({currency})"])
    for i, m in enumerate(revenue.get("monthly_revenue", [])):
        write_row(ws5, [m["month"], round(m["revenue"], 2)], alt=(i % 2 == 1))
    ws5.append([])
    write_header(ws5, ["Source", f"Revenue ({currency})"], fill=ACCENT_FILL)
    for i, (src, amt) in enumerate(revenue.get("revenue_by_source", {}).items()):
        write_row(ws5, [src, round(amt, 2)], alt=(i % 2 == 1))
    auto_width(ws5)

    # ---- Sheet 6: Platform Performance ----
    ws6 = wb.create_sheet("Platform Performance")
    add_title(ws6, "Platform Performance")
    write_header(ws6, ["Platform", "Views", "Reach", "Likes", "Comments", "Avg. Engagement (%)"])
    for i, p in enumerate(report_data.get("platform_performance", [])):
        write_row(ws6, [
            p.get("platform", ""),
            p.get("total_views", 0),
            p.get("total_reach", 0),
            p.get("total_likes", 0),
            p.get("total_comments", 0),
            round(p.get("average_engagement_rate", 0), 2),
        ], alt=(i % 2 == 1))
    auto_width(ws6)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
