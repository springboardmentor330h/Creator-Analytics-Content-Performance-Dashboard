from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import (
    getSampleStyleSheet,
    ParagraphStyle,
)
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)


# ============================================================
# PDF REPORT
# ============================================================

def generate_pdf_report(report_data: dict):
    """
    Generate a structured PDF analytics report.

    The function consumes the existing report_data generated
    by the reporting service. No analytics logic is duplicated.
    """

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40,
        title="CreatorIQ Analytics Report",
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontSize=20,
        leading=24,
        alignment=TA_CENTER,
        spaceAfter=12,
    )

    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        leading=14,
        alignment=TA_CENTER,
        spaceAfter=20,
    )

    heading_style = ParagraphStyle(
        "ReportHeading",
        parent=styles["Heading2"],
        fontSize=14,
        leading=18,
        spaceBefore=12,
        spaceAfter=8,
    )

    normal_style = ParagraphStyle(
        "ReportNormal",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
    )

    story = []

    creator_id = report_data.get("creator_id", "N/A")

    # --------------------------------------------------------
    # TITLE
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "CreatorIQ Analytics Report",
            title_style,
        )
    )

    story.append(
        Paragraph(
            f"Creator ID: {creator_id}",
            subtitle_style,
        )
    )

    # --------------------------------------------------------
    # HELPER FUNCTIONS
    # --------------------------------------------------------

    def money(value):
        try:
            return f"₹{float(value):,.2f}"
        except (TypeError, ValueError):
            return str(value)

    def add_section_title(title):
        story.append(
            Paragraph(
                title,
                heading_style,
            )
        )

    def add_simple_table(headers, rows):
        if not rows:
            story.append(
                Paragraph(
                    "No data available.",
                    normal_style,
                )
            )
            return

        table_data = [headers] + rows

        table = Table(
            table_data,
            repeatRows=1,
            hAlign="LEFT",
        )

        table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor("#1F4E78"),
                    ),
                    (
                        "TEXTCOLOR",
                        (0, 0),
                        (-1, 0),
                        colors.white,
                    ),
                    (
                        "FONTNAME",
                        (0, 0),
                        (-1, 0),
                        "Helvetica-Bold",
                    ),
                    (
                        "FONTNAME",
                        (0, 1),
                        (-1, -1),
                        "Helvetica",
                    ),
                    (
                        "FONTSIZE",
                        (0, 0),
                        (-1, -1),
                        8,
                    ),
                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.5,
                        colors.grey,
                    ),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [
                            colors.white,
                            colors.HexColor("#F3F6F9"),
                        ],
                    ),
                    (
                        "VALIGN",
                        (0, 0),
                        (-1, -1),
                        "MIDDLE",
                    ),
                    (
                        "LEFTPADDING",
                        (0, 0),
                        (-1, -1),
                        6,
                    ),
                    (
                        "RIGHTPADDING",
                        (0, 0),
                        (-1, -1),
                        6,
                    ),
                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        6,
                    ),
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        6,
                    ),
                ]
            )
        )

        story.append(table)
        story.append(Spacer(1, 12))

    # --------------------------------------------------------
    # REVENUE
    # --------------------------------------------------------

    revenue = report_data.get("revenue", {})

    if revenue:

        add_section_title("Revenue Analytics")

        revenue_summary = revenue.get(
            "revenue_summary",
            {},
        )

        total_revenue = revenue_summary.get(
            "total_revenue",
            0,
        )

        summary_rows = [
            [
                "Creator ID",
                str(
                    revenue_summary.get(
                        "creator_id",
                        creator_id,
                    )
                ),
            ],
            [
                "Total Revenue",
                money(total_revenue),
            ],
        ]

        add_simple_table(
            [
                "Metric",
                "Value",
            ],
            summary_rows,
        )

        # ----------------------------------------------------
        # REVENUE BY SOURCE
        # ----------------------------------------------------

        revenue_by_source = revenue.get(
            "revenue_by_source",
            {},
        )

        source_rows = revenue_by_source.get(
            "revenue_by_source",
            [],
        )

        if source_rows:

            add_section_title(
                "Revenue by Source"
            )

            rows = []

            for item in source_rows:

                if not isinstance(item, dict):
                    continue

                rows.append(
                    [
                        str(
                            item.get(
                                "source",
                                "Unknown",
                            )
                        ),
                        money(
                            item.get(
                                "amount",
                                0,
                            )
                        ),
                    ]
                )

            add_simple_table(
                [
                    "Revenue Source",
                    "Amount",
                ],
                rows,
            )

        # ----------------------------------------------------
        # MONTHLY REVENUE
        # ----------------------------------------------------

        monthly_revenue = revenue.get(
            "monthly_revenue",
            {},
        )

        monthly_rows = monthly_revenue.get(
            "monthly_revenue",
            [],
        )

        if monthly_rows:

            add_section_title(
                "Monthly Revenue"
            )

            rows = []

            for item in monthly_rows:

                if not isinstance(item, dict):
                    continue

                rows.append(
                    [
                        str(
                            item.get(
                                "month",
                                "Unknown",
                            )
                        ),
                        money(
                            item.get(
                                "amount",
                                0,
                            )
                        ),
                    ]
                )

            add_simple_table(
                [
                    "Month",
                    "Revenue",
                ],
                rows,
            )

    # --------------------------------------------------------
    # OTHER REPORT SECTIONS
    # --------------------------------------------------------

    def add_generic_section(
        section_name,
        section_data,
    ):
        if not section_data:
            return

        add_section_title(
            section_name
        )

        if isinstance(section_data, dict):

            rows = []

            for key, value in section_data.items():

                if isinstance(value, list):
                    continue

                if isinstance(value, dict):
                    continue

                rows.append(
                    [
                        str(key).replace(
                            "_",
                            " ",
                        ).title(),
                        str(value),
                    ]
                )

            if rows:
                add_simple_table(
                    [
                        "Metric",
                        "Value",
                    ],
                    rows,
                )

    # These sections can be populated by your
    # existing reporting service without changing
    # this export layer.

    add_generic_section(
        "Content Performance",
        report_data.get(
            "content_performance"
        ),
    )

    add_generic_section(
        "Audience Analytics",
        report_data.get(
            "audience_analytics"
        ),
    )

    add_generic_section(
        "Growth Trends",
        report_data.get(
            "growth_trends"
        ),
    )

    add_generic_section(
        "Platform Comparison",
        report_data.get(
            "platform_comparison"
        ),
    )

    # --------------------------------------------------------
    # BUILD PDF
    # --------------------------------------------------------

    document.build(story)

    buffer.seek(0)

    return buffer


# ============================================================
# EXCEL REPORT
# ============================================================

def generate_excel_report(report_data: dict):
    """
    Generate a structured Excel analytics report.

    Uses the existing report_data from the reporting service.
    """

    workbook = Workbook()

    worksheet = workbook.active

    if worksheet is None:
        raise RuntimeError(
            "Unable to access the active Excel worksheet."
        )

    worksheet.title = "Creator Report"

    # --------------------------------------------------------
    # STYLES
    # --------------------------------------------------------

    title_font = Font(
        bold=True,
        size=18,
    )

    section_font = Font(
        bold=True,
        size=13,
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    section_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    thin_border = Border(
        left=Side(style="thin", color="B7B7B7"),
        right=Side(style="thin", color="B7B7B7"),
        top=Side(style="thin", color="B7B7B7"),
        bottom=Side(style="thin", color="B7B7B7"),
    )

    # --------------------------------------------------------
    # HELPERS
    # --------------------------------------------------------

    row = 1

    def add_title(title):
        nonlocal row

        worksheet.cell(
            row=row,
            column=1,
            value=title,
        )

        worksheet.cell(
            row=row,
            column=1,
        ).font = title_font

        row += 1

    def add_section(title):
        nonlocal row

        worksheet.cell(
            row=row,
            column=1,
            value=title,
        )

        cell = worksheet.cell(
            row=row,
            column=1,
        )

        cell.font = section_font
        cell.fill = section_fill

        row += 1

    def add_table(headers, data_rows):
        nonlocal row

        if not data_rows:
            return

        # Header
        for column_index, header in enumerate(
            headers,
            start=1,
        ):

            cell = worksheet.cell(
                row=row,
                column=column_index,
                value=header,
            )

            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center"
            )

        row += 1

        # Data
        for data_row in data_rows:

            for column_index, value in enumerate(
                data_row,
                start=1,
            ):

                cell = worksheet.cell(
                    row=row,
                    column=column_index,
                    value=value,
                )

                cell.border = thin_border

                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'

            row += 1

        row += 1

    def money_value(value):
        try:
            return float(value)
        except (TypeError, ValueError):
            return value

    # --------------------------------------------------------
    # TITLE
    # --------------------------------------------------------

    add_title(
        "CreatorIQ Analytics Report"
    )

    worksheet.cell(
        row=row,
        column=1,
        value="Creator ID",
    )

    worksheet.cell(
        row=row,
        column=2,
        value=report_data.get(
            "creator_id",
            "N/A",
        ),
    )

    row += 2

    # --------------------------------------------------------
    # REVENUE
    # --------------------------------------------------------

    revenue = report_data.get(
        "revenue",
        {},
    )

    if revenue:

        add_section(
            "Revenue Summary"
        )

        revenue_summary = revenue.get(
            "revenue_summary",
            {},
        )

        total_revenue = revenue_summary.get(
            "total_revenue",
            0,
        )

        add_table(
            [
                "Metric",
                "Value",
            ],
            [
                [
                    "Creator ID",
                    revenue_summary.get(
                        "creator_id",
                        report_data.get(
                            "creator_id"
                        ),
                    ),
                ],
                [
                    "Total Revenue",
                    money_value(
                        total_revenue
                    ),
                ],
            ],
        )

        # ----------------------------------------------------
        # SOURCE
        # ----------------------------------------------------

        revenue_by_source = revenue.get(
            "revenue_by_source",
            {},
        )

        source_rows = revenue_by_source.get(
            "revenue_by_source",
            [],
        )

        if source_rows:

            add_section(
                "Revenue by Source"
            )

            rows = []

            for item in source_rows:

                if not isinstance(item, dict):
                    continue

                rows.append(
                    [
                        item.get(
                            "source",
                            "Unknown",
                        ),
                        money_value(
                            item.get(
                                "amount",
                                0,
                            )
                        ),
                    ]
                )

            add_table(
                [
                    "Revenue Source",
                    "Amount",
                ],
                rows,
            )

        # ----------------------------------------------------
        # MONTHLY
        # ----------------------------------------------------

        monthly_revenue = revenue.get(
            "monthly_revenue",
            {},
        )

        monthly_rows = monthly_revenue.get(
            "monthly_revenue",
            [],
        )

        if monthly_rows:

            add_section(
                "Monthly Revenue"
            )

            rows = []

            for item in monthly_rows:

                if not isinstance(item, dict):
                    continue

                rows.append(
                    [
                        item.get(
                            "month",
                            "Unknown",
                        ),
                        money_value(
                            item.get(
                                "amount",
                                0,
                            )
                        ),
                    ]
                )

            add_table(
                [
                    "Month",
                    "Revenue",
                ],
                rows,
            )

    # --------------------------------------------------------
    # OTHER ANALYTICS
    # --------------------------------------------------------

    def add_generic_excel_section(
        section_name,
        section_data,
    ):
        if not section_data:
            return

        if not isinstance(
            section_data,
            dict,
        ):
            return

        simple_rows = []

        for key, value in section_data.items():

            if isinstance(value, (dict, list)):
                continue

            simple_rows.append(
                [
                    str(key).replace(
                        "_",
                        " ",
                    ).title(),
                    value,
                ]
            )

        if simple_rows:

            add_section(
                section_name
            )

            add_table(
                [
                    "Metric",
                    "Value",
                ],
                simple_rows,
            )

    add_generic_excel_section(
        "Content Performance",
        report_data.get(
            "content_performance"
        ),
    )

    add_generic_excel_section(
        "Audience Analytics",
        report_data.get(
            "audience_analytics"
        ),
    )

    add_generic_excel_section(
        "Growth Trends",
        report_data.get(
            "growth_trends"
        ),
    )

    add_generic_excel_section(
        "Platform Comparison",
        report_data.get(
            "platform_comparison"
        ),
    )

    # --------------------------------------------------------
    # EXCEL FORMATTING
    # --------------------------------------------------------

    worksheet.freeze_panes = "A5"

    worksheet.column_dimensions["A"].width = 32
    worksheet.column_dimensions["B"].width = 24
    worksheet.column_dimensions["C"].width = 24
    worksheet.column_dimensions["D"].width = 24

    # Auto-adjust other populated columns.
    for column_cells in worksheet.columns:

        column_index = column_cells[0].column

        if column_index is None:
            continue

        column_letter = get_column_letter(
            column_index
        )

        current_width = (
            worksheet.column_dimensions[
                column_letter
            ].width
            or 0
        )

        max_length = current_width

        for cell in column_cells:

            if cell.value is not None:

                max_length = max(
                    max_length,
                    len(str(cell.value)),
                )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(max_length + 2, 12),
            45,
        )

    # --------------------------------------------------------
    # SAVE
    # --------------------------------------------------------

    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    return buffer





