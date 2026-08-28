from openpyxl import Workbook
from openpyxl.styles import Font


def create_excel_report(report_data: dict, file_path: str):

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Creator Report"

    # Title
    sheet["A1"] = (
        f"CreatorIQ Analytics Report - "
        f"Creator {report_data['creator_id']}"
    )
    sheet["A1"].font = Font(bold=True, size=16)

    row = 3

    sections = {
        "Content Performance": report_data["content_performance"],
        "Audience Analytics": report_data["audience_analytics"],
        "Revenue Analytics": report_data["revenue_analytics"],
        "Growth Trends": report_data["growth_trends"]
    }

    for section_name, data in sections.items():

        sheet.cell(row=row, column=1, value=section_name)
        sheet.cell(row=row, column=1).font = Font(
            bold=True,
            size=14
        )

        row += 1

        sheet.cell(row=row, column=1, value="Metric")
        sheet.cell(row=row, column=2, value="Value")

        sheet.cell(row=row, column=1).font = Font(bold=True)
        sheet.cell(row=row, column=2).font = Font(bold=True)

        row += 1

        for key, value in data.items():
            metric = key.replace("_", " ").title()

            sheet.cell(row=row, column=1, value=metric)
            sheet.cell(row=row, column=2, value=value)

            row += 1

        row += 2

    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 20

    workbook.save(file_path)