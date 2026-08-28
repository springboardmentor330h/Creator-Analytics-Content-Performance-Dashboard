import io
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from openpyxl import Workbook
from sqlalchemy.orm import Session
from app.models.content import Content
from app.models.audience import Audience
from app.models.revenue import RevenueRecord
from app.services import analytics_service, audience_service, revenue_service


def _content_rows(db: Session, creator_id: int):
    items = db.query(Content).filter(Content.creator_id == creator_id).all()
    header = ["Title", "Platform", "Views", "Likes", "Comments", "Reach"]
    rows = [[c.content_title, c.platform, c.views, c.likes, c.comments, c.reach] for c in items]
    return header, rows


def generate_content_pdf(db: Session, creator_id: int) -> io.BytesIO:
    header, rows = _content_rows(db, creator_id)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [Paragraph(f"Content Analytics Report — Creator {creator_id}", styles["Title"])]

    table_data = [header] + rows
    table = Table(table_data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_content_excel(db: Session, creator_id: int) -> io.BytesIO:
    header, rows = _content_rows(db, creator_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Content Analytics"
    ws.append(header)
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_audience_excel(db: Session, creator_id: int) -> io.BytesIO:
    items = db.query(Audience).filter(Audience.creator_id == creator_id).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Audience"
    ws.append(["Age Group", "Gender", "Country", "City", "Device", "Followers", "Reach"])
    for a in items:
        ws.append([a.age_group, a.gender, a.country, a.city, a.device_type, a.followers, a.reach])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_revenue_excel(db: Session, creator_id: int) -> io.BytesIO:
    items = db.query(RevenueRecord).filter(RevenueRecord.creator_id == creator_id).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue"
    ws.append(["Source", "Platform", "Amount", "Currency", "Date"])
    for r in items:
        ws.append([r.source, r.platform, r.amount, r.currency, r.earned_date.isoformat()])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_combined_report(db: Session, creator_id: int) -> dict:
    """
    Reuses existing services — no duplicate analytics logic per the
    sprint's explicit requirement.
    """
    return {
        "creator_id": creator_id,
        "content_summary": analytics_service.get_kpi_summary(db),
        "audience_summary": audience_service.get_audience_report(db),
        "revenue_summary": revenue_service.get_revenue_summary(db, creator_id),
        "growth_summary": {
            "monthly_revenue": revenue_service.get_monthly_revenue(db, creator_id),
            "trend": revenue_service.get_revenue_trend(db, creator_id),
        },
        "platform_comparison": analytics_service.get_platform_comparison(db),
    }