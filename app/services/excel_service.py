from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def generate_excel_report(report_data: dict):
    workbook = Workbook()

    # =====================================================
    # CONTENT PERFORMANCE SHEET
    # =====================================================

    content_sheet = workbook.active
    content_sheet.title = "Content Performance"

    content_headers = [
        "Content ID",
        "Platform",
        "Content Title",
        "Views",
        "Likes",
        "Comments",
        "Shares",
        "Saves",
        "Reach",
        "Total Engagement",
        "Engagement Rate"
    ]

    content_sheet.append(content_headers)

    for cell in content_sheet[1]:
        cell.font = Font(bold=True)

    for content in report_data.get("content_performance", []):
        content_sheet.append([
            content.get("content_id", 0),
            content.get("platform", ""),
            content.get("content_title", ""),
            content.get("views", 0),
            content.get("likes", 0),
            content.get("comments", 0),
            content.get("shares", 0),
            content.get("saves", 0),
            content.get("reach", 0),
            content.get("total_engagement", 0),
            content.get("engagement_rate", 0)
        ])

    if len(content_sheet["A"]) == 1:
        content_sheet.append(
            ["No data", "-", "-", 0, 0, 0, 0, 0, 0, 0, 0]
        )

    # =====================================================
    # AUDIENCE ANALYTICS SHEET
    # =====================================================

    audience_sheet = workbook.create_sheet("Audience Analytics")

    audience_headers = [
        "ID",
        "Age Group",
        "Gender",
        "Country",
        "City",
        "Device Type",
        "Active Hour",
        "Followers",
        "Impressions",
        "Reach"
    ]

    audience_sheet.append(audience_headers)

    for cell in audience_sheet[1]:
        cell.font = Font(bold=True)

    for audience in report_data.get("audience_analytics", []):
        audience_sheet.append([
            audience.get("id", 0),
            audience.get("age_group", ""),
            audience.get("gender", ""),
            audience.get("country", ""),
            audience.get("city", ""),
            audience.get("device_type", ""),
            audience.get("active_hour", 0),
            audience.get("followers", 0),
            audience.get("impressions", 0),
            audience.get("reach", 0)
        ])

    if len(audience_sheet["A"]) == 1:
        audience_sheet.append(
            ["No data", "-", "-", "-", "-", "-", 0, 0, 0, 0]
        )

    # =====================================================
    # GROWTH TRENDS SHEET
    # =====================================================

    growth_sheet = workbook.create_sheet("Growth Trends")

    growth_headers = [
        "Date",
        "Followers",
        "Reach",
        "Engagement Rate"
    ]

    growth_sheet.append(growth_headers)

    for cell in growth_sheet[1]:
        cell.font = Font(bold=True)

    for growth in report_data.get("growth_trends", []):
        growth_sheet.append([
            str(growth.get("date", "")),
            growth.get("followers", 0),
            growth.get("reach", 0),
            growth.get("engagement_rate", 0)
        ])

    if len(growth_sheet["A"]) == 1:
        growth_sheet.append(
            ["No data", 0, 0, 0]
        )

    # =====================================================
    # PLATFORM COMPARISON SHEET
    # =====================================================

    platform_sheet = workbook.create_sheet("Platform Comparison")

    platform_headers = [
        "Platform",
        "Content Count",
        "Views",
        "Reach",
        "Likes",
        "Comments",
        "Shares",
        "Total Engagement",
        "Average Engagement Rate"
    ]

    platform_sheet.append(platform_headers)

    for cell in platform_sheet[1]:
        cell.font = Font(bold=True)

    platform_data = report_data.get(
        "platform_comparison",
        []
    )

    # Handle dictionary format safely
    if isinstance(platform_data, dict):
        platform_data = list(platform_data.values())

    for platform in platform_data:
        platform_sheet.append([
            platform.get("platform", ""),
            platform.get("content_count", 0),
            platform.get("views", 0),
            platform.get("reach", 0),
            platform.get("likes", 0),
            platform.get("comments", 0),
            platform.get("shares", 0),
            platform.get("total_engagement", 0),
            platform.get("average_engagement_rate", 0)
        ])

    if len(platform_sheet["A"]) == 1:
        platform_sheet.append(
            ["No data", 0, 0, 0, 0, 0, 0, 0, 0]
        )

    # =====================================================
    # REVENUE ANALYTICS SHEET
    # =====================================================

    revenue_sheet = workbook.create_sheet("Revenue Analytics")

    revenue_sheet.append([
        "Metric",
        "Value"
    ])

    for cell in revenue_sheet[1]:
        cell.font = Font(bold=True)

    revenue = report_data.get(
        "revenue_analytics",
        {}
    )

    revenue_sheet.append([
        "Total Revenue",
        revenue.get("total_revenue", 0)
    ])

    for item in revenue.get("revenue_by_source", []):
        revenue_sheet.append([
            f"Revenue - {item.get('source', '')}",
            item.get("total_amount", 0)
        ])

    # =====================================================
    # AUTO COLUMN WIDTH
    # =====================================================

    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            sheet.column_dimensions[
                column_letter
            ].width = min(max_length + 2, 40)

    # =====================================================
    # SAVE TO MEMORY
    # =====================================================

    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    return buffer