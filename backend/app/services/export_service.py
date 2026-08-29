"""
export_service.py

Turns a report dict (produced by report_service.generate_creator_report)
into downloadable PDF or Excel bytes. Pure presentation layer — no
analytics logic lives here.
"""

import io
from typing import Any, Dict

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle


def generate_pdf_report(report: Dict[str, Any]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("CreatorIQ Performance Report", styles["Title"]))
    elements.append(Paragraph(f"Creator ID: {report['creator_id']}", styles["Normal"]))
    elements.append(Paragraph(f"Generated: {report['generated_at']}", styles["Normal"]))
    elements.append(Spacer(1, 16))

    # --- Content performance ---
    cp = report["content_performance"]
    elements.append(Paragraph("Content Performance Summary", styles["Heading2"]))
    data = [
        ["Metric", "Value"],
        ["Total Content", cp["total_content"]],
        ["Total Views", cp["total_views"]],
        ["Total Likes", cp["total_likes"]],
        ["Overall Engagement Rate (%)", cp["overall_engagement_rate"]],
    ]
    elements.append(_styled_table(data))
    elements.append(Spacer(1, 16))

    # --- Top content ---
    elements.append(Paragraph("Top Performing Content", styles["Heading2"]))
    data = [["Title", "Platform", "Views", "Engagement %"]]
    for item in report["top_content"]:
        data.append([
            item["content_title"][:45],
            item["platform"],
            item["views"],
            item["engagement_rate"],
        ])
    elements.append(_styled_table(data))
    elements.append(Spacer(1, 16))

    # --- Platform comparison ---
    elements.append(Paragraph("Platform Comparison", styles["Heading2"]))
    data = [["Platform", "Content Count", "Total Views", "Total Likes", "Avg Engagement %"]]
    for p in report["platform_comparison"]:
        data.append([p["platform"], p["content_count"], p["total_views"], p["total_likes"], p["avg_engagement_rate"]])
    elements.append(_styled_table(data))
    elements.append(Spacer(1, 16))

    # --- Revenue summary ---
    rev = report["revenue_analytics"]
    elements.append(Paragraph("Revenue Summary", styles["Heading2"]))
    data = [["Metric", "Value"], ["Total Revenue", rev["total_revenue"]]]
    for source, amount in rev["revenue_by_source"].items():
        data.append([f"Revenue — {source}", amount])
    elements.append(_styled_table(data))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def _styled_table(data: list) -> Table:
    table = Table(data, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    return table


def generate_excel_report(report: Dict[str, Any]) -> bytes:
    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    ws.append(["CreatorIQ Performance Report"])
    ws.append([f"Creator ID: {report['creator_id']}"])
    ws.append([f"Generated: {report['generated_at']}"])
    ws.append([])

    cp = report["content_performance"]
    ws.append(["Content Performance"])
    ws.append(["Total Content", cp["total_content"]])
    ws.append(["Total Views", cp["total_views"]])
    ws.append(["Total Likes", cp["total_likes"]])
    ws.append(["Overall Engagement Rate (%)", cp["overall_engagement_rate"]])
    ws.append([])

    rev = report["revenue_analytics"]
    ws.append(["Revenue Summary"])
    ws.append(["Total Revenue", rev["total_revenue"]])
    for source, amount in rev["revenue_by_source"].items():
        ws.append([f"Revenue — {source}", amount])

    # --- Top content sheet ---
    ws2 = wb.create_sheet("Top Content")
    ws2.append(["Title", "Platform", "Views", "Engagement %"])
    for item in report["top_content"]:
        ws2.append([item["content_title"], item["platform"], item["views"], item["engagement_rate"]])

    # --- Platform comparison sheet ---
    ws3 = wb.create_sheet("Platform Comparison")
    ws3.append(["Platform", "Content Count", "Total Views", "Total Likes", "Avg Engagement %"])
    for p in report["platform_comparison"]:
        ws3.append([p["platform"], p["content_count"], p["total_views"], p["total_likes"], p["avg_engagement_rate"]])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()